from fastapi import APIRouter, HTTPException, status, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime
import io
import openpyxl
from openpyxl import Workbook

from models import (
    Category, CategoryCreate, CategoryUpdate,
    Product, ProductCreate, ProductUpdate,
    Customer, CustomerCreate, CustomerUpdate,
    Sale, SaleCreate, SaleItem, SaleStatus,
    Quote, QuoteCreate, QuoteUpdate, QuoteStatus,
    UserRole
)
from permissions import Permission, require_permission
from database import (
    categories_collection, products_collection, customers_collection,
    sales_collection, quotes_collection, serialize_doc, serialize_docs,
    log_audit
)

# ===== CATEGORIES ROUTES =====

async def get_categories_list(current_user: dict):
    """List all categories"""
    categories = await categories_collection.find({}, {"_id": 0}).to_list(1000)
    return categories

async def create_category(category_data: CategoryCreate, current_user: dict):
    """Create new category"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_CREAR)
    
    # Check if category name already exists
    existing = await categories_collection.find_one({"nombre": category_data.nombre})
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ya existe una categoría con ese nombre"
        )
    
    category = Category(**category_data.model_dump())
    doc = category.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await categories_collection.insert_one(doc)
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="crear",
        modulo="categorias",
        detalles=f"Categoría creada: {category.nombre}"
    )
    
    return category

async def update_category(category_id: str, category_data: CategoryUpdate, current_user: dict):
    """Update category"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_EDITAR)
    
    existing = await categories_collection.find_one({"id": category_id})
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Categoría no encontrada"
        )
    
    update_data = category_data.model_dump(exclude_unset=True)
    
    if update_data:
        await categories_collection.update_one(
            {"id": category_id},
            {"$set": update_data}
        )
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="actualizar",
        modulo="categorias",
        detalles=f"Categoría actualizada: {category_id}"
    )
    
    updated = await categories_collection.find_one({"id": category_id}, {"_id": 0})
    return Category(**updated)

async def delete_category(category_id: str, current_user: dict):
    """Delete category"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_ELIMINAR)
    
    existing = await categories_collection.find_one({"id": category_id})
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Categoría no encontrada"
        )
    
    # Check if category is being used by products
    products_count = await products_collection.count_documents({"categoria_id": category_id})
    if products_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se puede eliminar. La categoría está siendo usada por {products_count} producto(s)"
        )
    
    await categories_collection.delete_one({"id": category_id})
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="eliminar",
        modulo="categorias",
        detalles=f"Categoría eliminada: {category_id}"
    )
    
    return {"message": "Categoría eliminada exitosamente"}

# ===== PRODUCTS ROUTES =====

async def get_products_list(current_user: dict):
    """List all products"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_VER)
    
    products = await products_collection.find({}, {"_id": 0}).to_list(1000)
    return products

async def search_products(query: str, current_user: dict):
    """Search products by name or barcode (autocomplete)"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_VER)
    
    # Search in name and barcode
    products = await products_collection.find({
        "$or": [
            {"nombre": {"$regex": query, "$options": "i"}},
            {"codigo_barras": {"$regex": query, "$options": "i"}}
        ]
    }, {"_id": 0}).limit(20).to_list(20)
    
    return products

async def get_product(product_id: str, current_user: dict):
    """Get product by ID"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_VER)
    
    product = await products_collection.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Producto no encontrado"
        )
    
    return Product(**product)

async def create_product(product_data: ProductCreate, current_user: dict):
    """Create new product"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_CREAR)
    
    # Check if barcode already exists
    if product_data.codigo_barras:
        existing = await products_collection.find_one({"codigo_barras": product_data.codigo_barras})
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un producto con ese código de barras"
            )
    
    product = Product(**product_data.model_dump())
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await products_collection.insert_one(doc)
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="crear",
        modulo="productos",
        detalles=f"Producto creado: {product.nombre}"
    )
    
    return product

async def update_product(product_id: str, product_data: ProductUpdate, current_user: dict):
    """Update product"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_EDITAR)
    
    existing = await products_collection.find_one({"id": product_id})
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Producto no encontrado"
        )
    
    update_data = product_data.model_dump(exclude_unset=True)
    
    # Check if barcode is being changed and already exists
    if "codigo_barras" in update_data and update_data["codigo_barras"]:
        barcode_exists = await products_collection.find_one({
            "codigo_barras": update_data["codigo_barras"],
            "id": {"$ne": product_id}
        })
        if barcode_exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un producto con ese código de barras"
            )
    
    update_data["updated_at"] = datetime.utcnow().isoformat()
    
    await products_collection.update_one(
        {"id": product_id},
        {"$set": update_data}
    )
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="actualizar",
        modulo="productos",
        detalles=f"Producto actualizado: {product_id}"
    )
    
    updated = await products_collection.find_one({"id": product_id}, {"_id": 0})
    return Product(**updated)

async def delete_product(product_id: str, current_user: dict):
    """Delete product"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_ELIMINAR)
    
    existing = await products_collection.find_one({"id": product_id})
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Producto no encontrado"
        )
    
    await products_collection.delete_one({"id": product_id})
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="eliminar",
        modulo="productos",
        detalles=f"Producto eliminado: {product_id}"
    )
    
    return {"message": "Producto eliminado exitosamente"}

# ===== CUSTOMERS ROUTES =====

async def get_customers_list(current_user: dict):
    """List all customers"""
    require_permission(UserRole(current_user["rol"]), Permission.CLIENTES_VER)
    
    customers = await customers_collection.find({}, {"_id": 0}).to_list(1000)
    return customers

async def get_customer(customer_id: str, current_user: dict):
    """Get customer by ID"""
    require_permission(UserRole(current_user["rol"]), Permission.CLIENTES_VER)
    
    customer = await customers_collection.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cliente no encontrado"
        )
    
    return Customer(**customer)

async def get_customer_history(customer_id: str, current_user: dict):
    """Get customer purchase history"""
    require_permission(UserRole(current_user["rol"]), Permission.CLIENTES_VER)
    
    customer = await customers_collection.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cliente no encontrado"
        )
    
    # Get all sales for this customer
    sales = await sales_collection.find(
        {"cliente_id": customer_id},
        {"_id": 0}
    ).sort("fecha", -1).to_list(100)
    
    total_compras = len(sales)
    total_gastado = sum(sale.get("total", 0) for sale in sales)
    
    return {
        "cliente": Customer(**customer),
        "total_compras": total_compras,
        "total_gastado": total_gastado,
        "ventas": sales
    }

async def create_customer(customer_data: CustomerCreate, current_user: dict):
    """Create new customer"""
    require_permission(UserRole(current_user["rol"]), Permission.CLIENTES_CREAR)
    
    # Check if email already exists
    if customer_data.email:
        existing = await customers_collection.find_one({"email": customer_data.email})
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un cliente con ese email"
            )
    
    customer = Customer(**customer_data.model_dump())
    doc = customer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await customers_collection.insert_one(doc)
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="crear",
        modulo="clientes",
        detalles=f"Cliente creado: {customer.nombre}"
    )
    
    return customer

async def update_customer(customer_id: str, customer_data: CustomerUpdate, current_user: dict):
    """Update customer"""
    require_permission(UserRole(current_user["rol"]), Permission.CLIENTES_EDITAR)
    
    existing = await customers_collection.find_one({"id": customer_id})
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cliente no encontrado"
        )
    
    update_data = customer_data.model_dump(exclude_unset=True)
    
    # Check if email is being changed and already exists
    if "email" in update_data and update_data["email"]:
        email_exists = await customers_collection.find_one({
            "email": update_data["email"],
            "id": {"$ne": customer_id}
        })
        if email_exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un cliente con ese email"
            )
    
    update_data["updated_at"] = datetime.utcnow().isoformat()
    
    await customers_collection.update_one(
        {"id": customer_id},
        {"$set": update_data}
    )
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="actualizar",
        modulo="clientes",
        detalles=f"Cliente actualizado: {customer_id}"
    )
    
    updated = await customers_collection.find_one({"id": customer_id}, {"_id": 0})
    return Customer(**updated)

async def delete_customer(customer_id: str, current_user: dict):
    """Delete customer"""
    require_permission(UserRole(current_user["rol"]), Permission.CLIENTES_ELIMINAR)
    
    existing = await customers_collection.find_one({"id": customer_id})
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cliente no encontrado"
        )
    
    await customers_collection.delete_one({"id": customer_id})
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="eliminar",
        modulo="clientes",
        detalles=f"Cliente eliminado: {customer_id}"
    )
    
    return {"message": "Cliente eliminado exitosamente"}

# ===== SALES ROUTES =====

async def get_sales_list(current_user: dict, limit: int = 100):
    """List all sales"""
    require_permission(UserRole(current_user["rol"]), Permission.VENTAS_VER)
    
    sales = await sales_collection.find({}, {"_id": 0}).sort("fecha", -1).limit(limit).to_list(limit)
    return sales

async def get_sale(sale_id: str, current_user: dict):
    """Get sale by ID"""
    require_permission(UserRole(current_user["rol"]), Permission.VENTAS_VER)
    
    sale = await sales_collection.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Venta no encontrada"
        )
    
    return Sale(**sale)

async def create_sale(sale_data: SaleCreate, current_user: dict):
    """Create new sale"""
    require_permission(UserRole(current_user["rol"]), Permission.VENTAS_CREAR)
    
    # Validate products exist and have stock
    for item in sale_data.items:
        product = await products_collection.find_one({"id": item.producto_id})
        if not product:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Producto no encontrado: {item.producto_id}"
            )
        
        if product["stock_actual"] < item.cantidad:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Stock insuficiente para {product['nombre']}. Stock actual: {product['stock_actual']}"
            )
    
    # Calculate total
    total = sum(item.subtotal for item in sale_data.items)
    
    # Create sale
    sale = Sale(
        **sale_data.model_dump(),
        vendedor_id=current_user["id"],
        vendedor_nombre=current_user["nombre"],
        total=total
    )
    
    doc = sale.model_dump()
    doc['fecha'] = doc['fecha'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await sales_collection.insert_one(doc)
    
    # Update product stock
    for item in sale_data.items:
        await products_collection.update_one(
            {"id": item.producto_id},
            {"$inc": {"stock_actual": -item.cantidad}}
        )
    
    # Update customer balance if applicable
    if sale_data.cliente_id:
        await customers_collection.update_one(
            {"id": sale_data.cliente_id},
            {"$inc": {"saldo_cuenta_corriente": total}}
        )
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="crear",
        modulo="ventas",
        detalles=f"Venta creada por ${total}"
    )
    
    return sale

async def anular_sale(sale_id: str, current_user: dict):
    """Anular (cancel) a sale"""
    require_permission(UserRole(current_user["rol"]), Permission.VENTAS_ANULAR)
    
    sale = await sales_collection.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Venta no encontrada"
        )
    
    if sale["estado"] == SaleStatus.ANULADA:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La venta ya está anulada"
        )
    
    # Restore product stock
    for item in sale["items"]:
        await products_collection.update_one(
            {"id": item["producto_id"]},
            {"$inc": {"stock_actual": item["cantidad"]}}
        )
    
    # Update customer balance if applicable
    if sale.get("cliente_id"):
        await customers_collection.update_one(
            {"id": sale["cliente_id"]},
            {"$inc": {"saldo_cuenta_corriente": -sale["total"]}}
        )
    
    # Update sale status
    await sales_collection.update_one(
        {"id": sale_id},
        {"$set": {"estado": SaleStatus.ANULADA}}
    )
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="anular",
        modulo="ventas",
        detalles=f"Venta anulada: {sale_id}"
    )
    
    return {"message": "Venta anulada exitosamente"}

# ===== QUOTES ROUTES =====

async def get_quotes_list(current_user: dict):
    """List all quotes"""
    require_permission(UserRole(current_user["rol"]), Permission.PRESUPUESTOS_VER)
    
    quotes = await quotes_collection.find({}, {"_id": 0}).sort("fecha", -1).to_list(100)
    return quotes

async def get_quote(quote_id: str, current_user: dict):
    """Get quote by ID"""
    require_permission(UserRole(current_user["rol"]), Permission.PRESUPUESTOS_VER)
    
    quote = await quotes_collection.find_one({"id": quote_id}, {"_id": 0})
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado"
        )
    
    return Quote(**quote)

async def create_quote(quote_data: QuoteCreate, current_user: dict):
    """Create new quote"""
    require_permission(UserRole(current_user["rol"]), Permission.PRESUPUESTOS_CREAR)
    
    # Validate products exist
    for item in quote_data.items:
        product = await products_collection.find_one({"id": item.producto_id})
        if not product:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Producto no encontrado: {item.producto_id}"
            )
    
    # Calculate total
    total = sum(item.subtotal for item in quote_data.items)
    
    # Create quote
    quote = Quote(
        **quote_data.model_dump(),
        vendedor_id=current_user["id"],
        vendedor_nombre=current_user["nombre"],
        total=total
    )
    
    doc = quote.model_dump()
    doc['fecha'] = doc['fecha'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await quotes_collection.insert_one(doc)
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="crear",
        modulo="presupuestos",
        detalles=f"Presupuesto creado por ${total}"
    )
    
    return quote

async def update_quote_status(quote_id: str, quote_data: QuoteUpdate, current_user: dict):
    """Update quote status"""
    require_permission(UserRole(current_user["rol"]), Permission.PRESUPUESTOS_EDITAR)
    
    quote = await quotes_collection.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado"
        )
    
    update_data = quote_data.model_dump(exclude_unset=True)
    
    await quotes_collection.update_one(
        {"id": quote_id},
        {"$set": update_data}
    )
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="actualizar",
        modulo="presupuestos",
        detalles=f"Presupuesto actualizado: {quote_id}"
    )
    
    updated = await quotes_collection.find_one({"id": quote_id}, {"_id": 0})
    return Quote(**updated)

async def convert_quote_to_sale(quote_id: str, current_user: dict):
    """Convert quote to sale"""
    require_permission(UserRole(current_user["rol"]), Permission.PRESUPUESTOS_CONVERTIR)
    
    quote = await quotes_collection.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado"
        )
    
    if quote["estado"] == QuoteStatus.CONVERTIDO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El presupuesto ya fue convertido a venta"
        )
    
    # Create sale from quote
    sale_items = [SaleItem(**item) for item in quote["items"]]
    
    sale_data = SaleCreate(
        cliente_id=quote["cliente_id"],
        items=sale_items,
        metodo_pago="efectivo",  # Default, can be changed
        notas=f"Convertido desde presupuesto {quote_id}"
    )
    
    # Validate stock
    for item in sale_items:
        product = await products_collection.find_one({"id": item.producto_id})
        if not product:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Producto no encontrado: {item.producto_id}"
            )
        
        if product["stock_actual"] < item.cantidad:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Stock insuficiente para {product['nombre']}. Stock actual: {product['stock_actual']}"
            )
    
    # Create sale
    sale = Sale(
        **sale_data.model_dump(),
        vendedor_id=current_user["id"],
        vendedor_nombre=current_user["nombre"],
        total=quote["total"]
    )
    
    sale_doc = sale.model_dump()
    sale_doc['fecha'] = sale_doc['fecha'].isoformat()
    sale_doc['created_at'] = sale_doc['created_at'].isoformat()
    
    await sales_collection.insert_one(sale_doc)
    
    # Update stock
    for item in sale_items:
        await products_collection.update_one(
            {"id": item.producto_id},
            {"$inc": {"stock_actual": -item.cantidad}}
        )
    
    # Update quote status
    await quotes_collection.update_one(
        {"id": quote_id},
        {"$set": {
            "estado": QuoteStatus.CONVERTIDO,
            "convertido_a_venta_id": sale.id
        }}
    )
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="convertir",
        modulo="presupuestos",
        detalles=f"Presupuesto {quote_id} convertido a venta {sale.id}"
    )
    
    return {
        "message": "Presupuesto convertido exitosamente",
        "venta": sale
    }

# ===== EXCEL ROUTES =====

async def import_products_excel(file: UploadFile, current_user: dict):
    """Import products from Excel file"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_CREAR)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser un Excel (.xlsx o .xls)"
        )
    
    try:
        # Read Excel file
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        products_created = 0
        products_updated = 0
        errors = []
        
        # Expected columns: nombre, descripcion, codigo_barras, precio_minorista, precio_mayorista, stock_actual, stock_minimo
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Skip empty rows
                    continue
                
                nombre = str(row[0]).strip()
                descripcion = str(row[1]).strip() if row[1] else None
                codigo_barras = str(row[2]).strip() if row[2] else None
                precio_minorista = float(row[3]) if row[3] else 0.0
                precio_mayorista = float(row[4]) if row[4] else 0.0
                stock_actual = int(row[5]) if row[5] else 0
                stock_minimo = int(row[6]) if row[6] else 0
                
                # Check if product exists by barcode
                existing = None
                if codigo_barras:
                    existing = await products_collection.find_one({"codigo_barras": codigo_barras})
                
                if existing:
                    # Update existing product
                    await products_collection.update_one(
                        {"id": existing["id"]},
                        {"$set": {
                            "nombre": nombre,
                            "descripcion": descripcion,
                            "precio_minorista": precio_minorista,
                            "precio_mayorista": precio_mayorista,
                            "stock_actual": stock_actual,
                            "stock_minimo": stock_minimo,
                            "updated_at": datetime.utcnow().isoformat()
                        }}
                    )
                    products_updated += 1
                else:
                    # Create new product
                    product = Product(
                        nombre=nombre,
                        descripcion=descripcion,
                        codigo_barras=codigo_barras,
                        precio_minorista=precio_minorista,
                        precio_mayorista=precio_mayorista,
                        stock_actual=stock_actual,
                        stock_minimo=stock_minimo
                    )
                    
                    doc = product.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    doc['updated_at'] = doc['updated_at'].isoformat()
                    
                    await products_collection.insert_one(doc)
                    products_created += 1
                    
            except Exception as e:
                errors.append(f"Fila {row_idx}: {str(e)}")
        
        await log_audit(
            usuario_id=current_user["id"],
            usuario_nombre=current_user["nombre"],
            accion="importar",
            modulo="productos",
            detalles=f"Excel importado: {products_created} creados, {products_updated} actualizados"
        )
        
        return {
            "message": "Importación completada",
            "productos_creados": products_created,
            "productos_actualizados": products_updated,
            "errores": errors
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error al procesar el archivo: {str(e)}"
        )

async def export_products_excel(current_user: dict):
    """Export products to Excel file"""
    require_permission(UserRole(current_user["rol"]), Permission.PRODUCTOS_VER)
    
    # Get all products
    products = await products_collection.find({}, {"_id": 0}).to_list(1000)
    
    # Create workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Productos"
    
    # Headers
    headers = ["Nombre", "Descripción", "Código Barras", "Precio Minorista", "Precio Mayorista", 
               "Stock Actual", "Stock Mínimo", "Categoría ID"]
    sheet.append(headers)
    
    # Data
    for product in products:
        sheet.append([
            product.get("nombre", ""),
            product.get("descripcion", ""),
            product.get("codigo_barras", ""),
            product.get("precio_minorista", 0),
            product.get("precio_mayorista", 0),
            product.get("stock_actual", 0),
            product.get("stock_minimo", 0),
            product.get("categoria_id", "")
        ])
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="exportar",
        modulo="productos",
        detalles=f"Excel exportado con {len(products)} productos"
    )
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"}
    )

# ===== BACKUP ROUTES =====

async def create_full_backup(current_user: dict):
    """Create full database backup"""
    from permissions import Permission
    from models import UserRole
    
    # Only admins can create backups
    require_permission(UserRole(current_user["rol"]), Permission.USUARIOS_VER)
    
    from database import (
        products_collection,
        categories_collection,
        customers_collection,
        sales_collection,
        quotes_collection,
        users_collection,
    )
    
    # Get all collections data
    backup_data = {
        "backup_date": datetime.utcnow().isoformat(),
        "backup_by": current_user["nombre"],
        "version": "1.0.0",
        "data": {
            "products": await products_collection.find({}, {"_id": 0}).to_list(10000),
            "categories": await categories_collection.find({}, {"_id": 0}).to_list(10000),
            "customers": await customers_collection.find({}, {"_id": 0}).to_list(10000),
            "sales": await sales_collection.find({}, {"_id": 0}).to_list(10000),
            "quotes": await quotes_collection.find({}, {"_id": 0}).to_list(10000),
            "users": await users_collection.find({}, {"_id": 0, "password_hash": 0}).to_list(10000),
        }
    }
    
    # Convert to JSON
    import json
    backup_json = json.dumps(backup_data, indent=2, ensure_ascii=False, default=str)
    
    await log_audit(
        usuario_id=current_user["id"],
        usuario_nombre=current_user["nombre"],
        accion="backup",
        modulo="sistema",
        detalles="Backup completo de la base de datos"
    )
    
    # Return as downloadable file
    return StreamingResponse(
        io.BytesIO(backup_json.encode('utf-8')),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=backup_ferreteria_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"}
    )
